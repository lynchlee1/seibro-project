import json
import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options

from settings import get

from openpyxl import load_workbook

def fmtkey(key):
    key=str(key).replace(' ','')
    types_str = [
        'EB', 'eb',
        'CB', 'cb',
        'BW', 'bw',
    ]
    for abbr in types_str: key=key.replace(abbr,'')
    idx=key.find('(')
    if idx!=-1: key=key[:idx]
    return key

def load_database():
    db_path = "database.json"
    try:
        with open(db_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return {}

def save_database(data, corp_name, last_modified, mode):
    db_path = "database.json"
    try:
        with open(db_path, 'r', encoding='utf-8') as f:
            db = json.load(f)
    except:
        db = {}

    db["last_modified"] = last_modified
    if corp_name not in db:
        db[corp_name] = {
            "hist": [],
            "prc": []
        }    
    db[corp_name][mode].extend(data)
    with open(db_path, 'w', encoding='utf-8') as f:
        json.dump(db, f, ensure_ascii=False, indent=2)

def read_companies_from_excel():
    wb = load_workbook('results.xlsx', data_only=True)
    sheet = wb['DB']
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    name_idx, round_idx = headers.index('기업명'), headers.index('회차')
    if name_idx is None or round_idx is None:
        print("Missing required columns: 기업명, 회차")
        return []

    targets = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        company = (row[name_idx] or '').strip() if isinstance(row[name_idx], str) else str(row[name_idx] or '').strip()
        round_num = (row[round_idx] or '').strip() if isinstance(row[round_idx], str) else str(row[round_idx] or '').strip()
        if not company or not round_num:
            continue
        targets.append({"company": company, "keyword": f"{company}{round_num}"})
    return targets

def scrape_once(config):
    print(f"\nStarting scrape for company: {config['company']}")
    print(f"Search keyword: {config['keyword']}")
    try:
        print("Initializing scraper...")
        scraper = KINDScraper(config, headless=False, process_type='hist')
        scraper.setup()
        print("Opening details page...")
        scraper.driver.get(get("details_url"))
        time.sleep(get("buffer_time"))
        print("Clicking search button...")
        scraper._click_button(get("search_button_selector"))
        time.sleep(get("buffer_time"))
        print("Filling company name...")
        scraper._fill_input(get("company_input_selector"), config["company"], in_iframe=True)
        scraper._click_button(get("company_search_selector"), in_iframe=True)
        time.sleep(get("buffer_time"))
        print("Switching to popup iframe...")
        iframe_selector = get("popup_iframe")
        iframe = scraper.driver.find_element(By.CSS_SELECTOR, iframe_selector)
        scraper.driver.switch_to.frame(iframe)
        container = scraper.driver.find_element(By.CSS_SELECTOR, "#group74 #group118 #isinList")
        items = container.find_elements(By.CSS_SELECTOR, '[id^="isinList_"][id$="_group178"]')
        searched_keys = []
        print("Searching for matching items...")
        for _, item in enumerate(items):
            text = scraper.driver.execute_script("return arguments[0].textContent.trim();", item) or ""
            searched_keys.append(fmtkey(text))
        pos = [i for i in range(len(searched_keys)) if searched_keys[i] == config.get("keyword")]
        if not pos:
            print("No matches found")
            return []
        target = pos[0]
        print(f"Found match at position {target}")
        scraper._click_button(f"#isinList_{target}_ISIN_ROW")
        scraper.driver.switch_to.default_content()
        print("Filling dates...")
        scraper._fill_dates()
        time.sleep(get("buffer_time"))
        print("Starting data collection...")
        scraper._click_button("#image2")
        time.sleep(get("short_loadtime"))

        all_rows_dicts = []
        previous_page_key = None
        page_num = 1
        while True:
            try:
                print(f"\nProcessing page {page_num}...")
                tbody = scraper.driver.find_element(By.CSS_SELECTOR, "#grid1_body_tbody")
                rows = tbody.find_elements(By.TAG_NAME, "tr")
                headers = []
                header_cells = []
                if rows:
                    header_cells = rows[0].find_elements(By.TAG_NAME, "th") or rows[0].find_elements(By.TAG_NAME, "td")
                    headers = scraper.driver.execute_script(
                        """
                        var result = [];
                        for (var i = 0; i < arguments[0].length; i++) {
                            var v = arguments[0][i].textContent.trim();
                            result.push(v || `col_${i}`);
                        }
                        return result;
                        """,
                        header_cells
                    )
                page_key = None
                if rows:
                    first_row_cells = rows[0].find_elements(By.TAG_NAME, "td") or rows[0].find_elements(By.TAG_NAME, "th")
                    first_vals = scraper.driver.execute_script(
                        """
                        var result = [];
                        for (var i = 0; i < arguments[0].length; i++) {
                            result.push(arguments[0][i].textContent.trim());
                        }
                        return result;
                        """,
                        first_row_cells
                    )
                    page_key = "|".join(first_vals)
                if previous_page_key is not None and page_key == previous_page_key:
                    print("Reached duplicate page, stopping")
                    break
                data_dicts = []
                start_idx = 1 if header_cells else 0
                for r_idx in range(start_idx, len(rows)):
                    row = rows[r_idx]
                    cells = row.find_elements(By.TAG_NAME, "td") or row.find_elements(By.TAG_NAME, "th")
                    if not cells:
                        continue
                    values = scraper.driver.execute_script(
                        """
                        var result = [];
                        for (var i = 0; i < arguments[0].length; i++) {
                            result.push(arguments[0][i].textContent.trim());
                        }
                        return result;
                        """,
                        cells
                    )
                    row_dict = {}
                    try:
                        row_dict["title"] = values[2]
                        row_dict["exc_start"] = values[3]
                        row_dict["exc_end"] = values[4]
                        row_dict["date"] = values[5]
                        row_dict["exc_amount"] = values[6]
                        row_dict["exc_shares"] = values[8]
                        row_dict["exc_price"] = values[9]
                        row_dict["listing_date"] = values[10]
                    except Exception:
                        pass
                    data_dicts.append(row_dict)
                all_rows_dicts.extend(data_dicts)
                previous_page_key = page_key
                try:
                    scraper._click_button("#gridPaging_next_btn")
                    time.sleep(get("short_loadtime"))
                    page_num += 1
                except Exception:
                    print("No more pages")
                    break
            except Exception:
                break
        return all_rows_dicts
    finally:
        try:
            scraper.cleanup()
        except Exception:
            pass

def run_scraper(from_date, to_date):
    targets = read_companies_from_excel()
    if not targets:
        print("No targets in results.xlsx; running single example config")
        
    results = []
    for i, t in enumerate(targets):
        print(f"\n=== [{i+1}/{len(targets)}] {t['company']} - {t['keyword']} ===")
        cfg = {
            'from_date': from_date,
            'to_date': to_date,
            'company': t['company'],
            'keyword': t['keyword']
        }
        try:
            rows = scrape_once(cfg) or []
            results.extend(rows)
            corp_name = t['company']
            save_database(results, corp_name, to_date, 'hist')
            print(f"Saved {len(results)} rows to database.json")
        except Exception as e:
            print(f"Skip {t['company']}: {e}")
            continue
    print("\n=== Process completed ===")

class KINDScraper:
    def __init__(self, config, headless = False, process_type = None):
        self.config = config # from_date, to_date, company, key
        self.headless = headless
        self.driver = None
        self.wait = None

    def setup(self): 
        self.driver, self.wait = self._setup_driver(headless=self.headless)

    def cleanup(self): 
        if self.driver: self.driver.quit()
    
    def _click_button(self, selector, in_iframe=False):
        try:
            # Switch to iframe if needed
            if in_iframe:
                iframe_selector = get("popup_iframe")
                iframe = self.driver.find_element(By.CSS_SELECTOR, iframe_selector)
                self.driver.switch_to.frame(iframe)
                print(f"🔄 Switched to iframe: {iframe_selector}")
            
            button = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
            self.driver.execute_script("arguments[0].click();", button)
            print(f"✅ CSS {selector} 버튼 클릭 완료")
            time.sleep(get("buffer_time"))
            
            # Switch back to main frame if we switched to iframe
            if in_iframe:
                self.driver.switch_to.default_content()
                print(f"🔄 Switched back to main frame")
                
        except Exception as e: 
            # Make sure to switch back to main frame on error
            if in_iframe:
                try:
                    self.driver.switch_to.default_content()
                except:
                    pass
            raise Exception(f"❌ CSS {selector} 클릭 실패: {e}")

    def _fill_input(self, selector, value, in_iframe=False):
        if value is None: return
        try:
            # Switch to iframe if needed
            if in_iframe:
                iframe_selector = get("popup_iframe")
                iframe = self.driver.find_element(By.CSS_SELECTOR, iframe_selector)
                self.driver.switch_to.frame(iframe)
                print(f"🔄 Switched to iframe: {iframe_selector}")
            
            # Check if selector is XPath or CSS
            if selector.startswith("//") or selector.startswith(".//"):
                input = self.driver.find_element(By.XPATH, selector)
                print(f"✅ XPATH {selector} 입력 완료")
            else:
                input = self.driver.find_element(By.CSS_SELECTOR, selector)
                print(f"✅ CSS {selector} 입력 완료")
            
            # Use JavaScript click to bypass popup overlay
            self.driver.execute_script("arguments[0].click();", input)
            time.sleep(0.5)
            input.clear()
            input.send_keys(value)
            time.sleep(get("buffer_time"))
            
            # Switch back to main frame if we switched to iframe
            if in_iframe:
                self.driver.switch_to.default_content()
                print(f"🔄 Switched back to main frame")
                
        except Exception as e: 
            # Make sure to switch back to main frame on error
            if in_iframe:
                try:
                    self.driver.switch_to.default_content()
                except:
                    pass
            raise Exception(f"❌ {selector} 입력 실패: {e}")
    
    def _fill_dates(self):
        try:
            self._fill_input(get("from_date_selector"), self.config.get('from_date'))
            self._fill_input(get("to_date_selector"), self.config.get('to_date'))
        except Exception as e: raise Exception(f"❌ 날짜 입력 실패: {e}")

    def _setup_driver(self, headless):
        chrome_options = Options()
        if headless:
            chrome_options.add_argument("--headless=new")
        
        # Stability and crash prevention options
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--disable-web-security")
        chrome_options.add_argument("--disable-features=VizDisplayCompositor")
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--disable-plugins")
        chrome_options.add_argument("--disable-images")
        chrome_options.add_argument("--disable-default-apps")
        chrome_options.add_argument("--disable-sync")
        chrome_options.add_argument("--disable-translate")
        
        # Memory and performance options
        chrome_options.add_argument("--memory-pressure-off")
        chrome_options.add_argument("--max_old_space_size=4096")
        chrome_options.add_argument("--disable-background-networking")
        chrome_options.add_argument("--disable-background-timer-throttling")
        chrome_options.add_argument("--disable-backgrounding-occluded-windows")
        chrome_options.add_argument("--disable-renderer-backgrounding")
        
        # Logging and notifications
        chrome_options.add_argument("--disable-notifications")
        chrome_options.add_argument("--disable-logging")
        chrome_options.add_argument("--log-level=3")
        chrome_options.add_argument("--silent")
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        chrome_options.add_argument("--window-size=1600,1000")

        # Use Selenium Manager (built into Selenium 4.6+) for consistent driver resolution
        print("Chrome driver 시작 중... (Selenium Manager)")
        driver = webdriver.Chrome(options=chrome_options)
        driver.set_page_load_timeout(get("long_loadtime"))
        wait = WebDriverWait(driver, get("long_loadtime"))
        print("✅ Chrome driver 로딩 완료")
        return driver, wait

    def run(self):
        pass

if __name__ == "__main__":
    run_scraper(from_date="20240101", to_date="20241231")